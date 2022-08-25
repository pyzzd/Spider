import re
import time
import openpyxl
import requests
import execjs
import hashlib

def hash(value: str, _mode: str) -> str:
    _hash = eval(f"hashlib.{_mode}(value.encode('utf-8')).hexdigest()")
    return _hash


def get_cookie(item: dict) -> str:
    ct = item.get('ct', None)
    bts = item.get('bts', None)
    chars = item.get('chars', None)
    hash_mode = item.get('ha', None)
    chars_length = len(item.get('chars', None))
    for i in range(chars_length):
        for j in range(chars_length):
            value = bts[0] + chars[i] + chars[j] + bts[1]
            if hash(value, hash_mode) == ct:
                return '__jsl_clearance=' + value


def update_headers(url):
    res = requests.get(url, headers=headers)
    if res.status_code == 200:
        return True
    __jsluid_h = '__jsluid_h=' + res.cookies['__jsluid_h']
    js1_text = re.findall('cookie=(.*?);location', res.text)[0]
    __jsl_clearance = str(execjs.eval(js1_text)).replace('max-age=3600;path=/', '')
    headers['Cookie'] = __jsluid_h + ';' + __jsl_clearance
    res1 = requests.get(url, headers=headers).text
    dict_text = re.findall(r';go\((.*?)\)', res1)[0]
    value = get_cookie(eval(dict_text))
    headers['Cookie'] = __jsluid_h + ';' + value + ';'

def prase():
    with open('info.txt','r',encoding='utf-8')as f:
        info_list = f.readlines()
    with open('已采集.txt','r',encoding='utf-8')as f:
        ycj_list = f.readlines()
    for info in info_list:
        href = info.split('   ')[2].replace('\n','')
        title = info.split('   ')[0]
        ptime = info.split('   ')[1]
        tp = info.split('   ')[3]
        print(href)
        if href in ycj_list:
            print('已采集')
            continue
        source = requests.get(href,headers=headers).text
        if '<script>document.cookie=' in source:
            update_headers(href)
        headers['Cookie']+='clientlogin=5187hOM3T9_QvnztjTNy1z7MMy9jGLVVw5odk.v1jp9V4QUKtFEqUcYctVMSKoA_hfXkBiEfVZHTcYnxqxx2jWt0IB652oKxnOfySTsDPQvJXCjRRCK9;'
        source = requests.get(href, headers=headers).text
        if '<span class="open_quick_reg">（略）</span>' in source:
            print('登录Cookie失效')
            quit()
        try:
            cjj = re.findall(r'成交价:(.*?)</',source,re.S)[0]
        except:
            cjj = re.findall('中标（成交）金额：(.*?)</',source,re.S)
            if cjj == []:
                try:
                    cjj = re.findall('合同金额：(.*?)</',source,re.S)[0]
                except:
                    try:
                        cjj = re.findall('成交金额：(.*?)</p>',source,re.S)[0]
                    except:
                        try:
                            cjj = re.findall(r'总中标金额：</span>(.*?)</span></p>',source,re.S)[0]
                        except:
                            cjj = ''
            else:
                cjj = ','.join(list(set(cjj)))

        try:
            gys = re.findall(r'<div>成交供应商:(.*?)</',source,re.S)[0]
        except:
            gys = re.findall('供应商名称：(.*?)</',source,re.S)
            if gys == []:
                try:
                    gys = re.findall('供应商（乙方）：(.*?)</',source,re.S)[0]
                except:
                    try:
                        gys = re.findall('成交供应商：(.*?)</p>',source,re.S)[0]
                    except:
                        try:
                            gys = re.findall('成交供应商:(.*?)</p>',source,re.S)[0]
                        except:
                            try:
                                gys = re.findall('中标人名称：(.*?)</span></p>',source,re.S)[0]
                            except:
                                gys = ''
            else:
                gys = ','.join(list(set(gys)))

        try:
            cgr = re.findall(r'采购单位：(.*?)</',source,re.S)[0]
        except:
            try:
                cgr = re.findall('采购人信息.*?名 称：(.*?)</p>',source,re.S)[0]
            except:
                try:
                    cgr = re.findall('采购人信息</span>.*?名称：.*?<.*?>(.*?)</', source,re.S)[0]
                except:
                    try:
                        cgr = re.findall('采购人（甲方）：(.*?)</',source,re.S)[0]
                    except:
                        try:
                            cgr = re.findall(r'采购人名称：(.*?)</p>',source,re.S)[0]
                        except:
                            cgr = ''
        try:
            dz = re.findall('送货地点：(.*?)</',source,re.S)[0]
        except:
            try:
                dz = re.findall(r'采购人信息</p>.*?地址：(.*?)</p>',source,re.S)[0]
            except:
                try:
                    dz = re.findall(r'采购人信息</span>.*?地址：.*?<.*?>(.*?)</', source,re.S)[0]
                except:
                    try:
                        dz = re.findall(r'地\s\s址：(.*?)</',source,re.S)[0]
                    except:
                        try:
                            dz = re.findall(r'采购人信息.*?地址：(.*?)</p>',source,re.S)[0]
                        except:
                            try:
                                dz = re.findall('采购人地址：(.*?)</p>',source,re.S)[0]
                            except:
                                dz = ''
        re_bq = re.compile('<.*?>')
        title = re_bq.sub('', title)
        cjj = re_bq.sub('', cjj)
        gys = re_bq.sub('', gys)
        cgr = re_bq.sub('', cgr)
        dz = re_bq.sub('', dz)
        cjj = str(cjj).strip()
        gys = str(gys).strip()
        cgr = str(cgr).strip()
        dz = str(dz).strip()
        time.sleep(1.5)
        with open('已采集.txt','a',encoding='utf-8')as f:
            f.write(href+'\n')
        if cjj == '' or gys == '' or cgr == '' or dz == '':
            print('未匹配到数据')
            continue
        save(title,ptime,cjj,gys,cgr,dz,href,tp)

def save(title,ptime,cjj,gys,cgr,dz,href,tp):
    global row
    print(row)
    print(title,ptime,cjj,gys,cgr,dz,tp)
    ws.cell(row,1,title)
    ws.cell(row, 2, ptime)
    ws.cell(row, 3, cjj)
    ws.cell(row, 4, gys)
    ws.cell(row, 5, cgr)
    ws.cell(row, 6, dz)
    ws.cell(row, 7, href)
    ws.cell(row,8,tp)
    row +=1
    wb.save('info.xlsx')
if __name__ == '__main__':
    url_list = []
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
        'Cookie': ''
    }
    wb = openpyxl.load_workbook('info.xlsx')
    ws = wb.active
    row = ws.max_row+1
    prase()
