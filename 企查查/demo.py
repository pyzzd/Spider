import re
import random
import time
import openpyxl
import requests
import hmac
import hashlib
import os


class Spider:
    def __init__(self):
        if os.path.exists('over.txt'):
            pass
        else:
            f = open('over.txt','w').close()
        self.normal_headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'zh-CN,zh;q=0.9',
            'sec-ch-ua': '".Not/A)Brand";v="99", "Google Chrome";v="103", "Chromium";v="103"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'sec-fetch-user': '?1',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
        }
        self.cookie_pool = []
        self.pool_len = len(self.cookie_pool)
        self.wb = openpyxl.load_workbook('info.xlsx')
        self.ws = self.wb.active
        self.row = self.ws.max_row+1
        self.row_ = self.ws.max_row+1
        self.max_coon = 3
        self.c = random.randint(0,2)

    def load_proxy(self):
        proxyHost = "http-dyn.abuyun.com"
        proxyPort = "9020"

        # 代理隧道验证信息
        proxyUser = "HJW47FL79000Z5ID"
        proxyPass = "0A2B2CCA9D5CB4DC"

        proxyMeta = "http://%(user)s:%(pass)s@%(host)s:%(port)s" % {
            "host": proxyHost,
            "port": proxyPort,
            "user": proxyUser,
            "pass": proxyPass,
        }
        self.proxy = None

    def load_no(self):
        with open('test.txt', 'r', encoding='utf-8') as f:
            no_list = f.readlines()
        return no_list

    def encrypt_to_hmac(self, key, value, type_=hashlib.sha512):
        """
        key:密钥key
        value:待加密的字符串
        type_:hash函数
        return: 加密后的16进制
        """
        mac = hmac.new(key.encode(encoding="utf-8"), value.encode("utf8"), type_)
        return mac.hexdigest()


    def res_zlxx(self,no):
        page = 1
        xh = 1
        while True:
            if page>10:
                with open('未采集.txt','a',encoding='utf-8')as f:
                    f.write(no+'\n')
                with open('over.txt','a',encoding='utf-8')as f:
                    f.write(no+'\n')
                return False
            print(str(self.c)+';'+f'正在采集专利信息第{page}页')
            now_res = 1
            if self.c>=self.pool_len:
                self.c = 0
            payload = {"keyNo": no, "pageIndex": page, "sortField": "applicationdate"}
            key = 'iLAgiWL4Ligk4ikQLLvigigk4iLAgiWL4Ligk4ikQLLvigigk4'
            header_name = '/api/datalist/zhuanlilist{}'.format(str(payload)).replace('N', 'n').replace('I', 'i').replace(
                'F', 'f').replace("'", '"').replace(' ', '')
            self.normal_headers['Cookie'] = self.cookie_pool[self.c]
            header_value = '/api/datalist/zhuanlilistpathString{}{}'.format(str(payload),self.tid[self.c]).replace('N', 'n').replace('I', 'i').replace('F', 'f').replace("'", '"').replace(' ', '')
            self.normal_headers[self.encrypt_to_hmac(key, header_name)[8:28]] = self.encrypt_to_hmac(key, header_value)
            res = ''
            while now_res <self.max_coon:
                try:
                    res = requests.post('https://www.qcc.com/api/datalist/zhuanlilist', json=payload,headers=self.normal_headers,timeout=3).json()
                    now_res = 3
                except:
                    if self.c == 3:
                        self.c = 0
                    else:
                        self.c +=1
                    self.normal_headers['Cookie'] = self.cookie_pool[self.c]
                    print('重连')
                    now_res+=1
            if res == '':
                pass
            if res == None:
                time.sleep(3.5)
                print('没有专利信息')
                break
            if res['data'] == []:
                print('获取完毕')
                break
            for d in res['data']:
                Title = d['Title']
                KindCodeDesc = d['KindCodeDesc']
                LegalStatusDesc = d['LegalStatusDesc']
                ApplicationNumber = d['ApplicationNumber']
                ApplicationDate = d['ApplicationDate']
                PublicationNumber = d['PublicationNumber']
                PublicationDate = d['PublicationDate']
                InventorList = d['InventorList']
                self.save_zlxx(xh,Title,KindCodeDesc,LegalStatusDesc,ApplicationNumber,ApplicationDate,PublicationNumber,PublicationDate,InventorList)
                xh+=1
            time.sleep(1)
            page+=1
            self.c += 1
        return True

    def res_gxy(self,no):
        page = 1

        while True:
            if page > 10:
                with open('未采集.txt','a',encoding='utf-8')as f:
                    f.write(no+'\n')
                with open('over.txt','a',encoding='utf-8')as f:
                    f.write(no+'\n')
                return False
            now_res = 1
            print(str(self.c)+';'+f'正在采集供应商信息第{page}页')
            if self.c >= self.pool_len:
                self.c = 0
            data = 'datatype=0&keyno={}&pageindex={}'.format(no,page)
            arguments = ["/api/datalist/suppliercustomer?{}".format(data)]
            i = {
                "default": {
                    "n": 20,
                    "codes": {
                        "0": "W",
                        "1": "l",
                        "2": "k",
                        "3": "B",
                        "4": "Q",
                        "5": "g",
                        "6": "f",
                        "7": "i",
                        "8": "i",
                        "9": "r",
                        "10": "v",
                        "11": "6",
                        "12": "A",
                        "13": "K",
                        "14": "N",
                        "15": "k",
                        "16": "4",
                        "17": "L",
                        "18": "1",
                        "19": "8",
                    }
                }
            }
            e = arguments[0]
            t = e + e
            key = ''
            for n in range(len(t)):
                s = ord(t[n]) % 20
                key += i['default']['codes'][str(s)]
            header_name = "/api/datalist/suppliercustomer?" + data + "{}"
            self.normal_headers['Cookie'] = self.cookie_pool[self.c]
            header_value = "/api/datalist/suppliercustomer?" + data + "pathString{}"+self.tid[self.c]
            self.normal_headers[self.encrypt_to_hmac(key, header_name)[8:28]] = self.encrypt_to_hmac(key, header_value)
            res = ''
            while now_res<self.max_coon:
                try:
                    res = requests.get('https://www.qcc.com/api/datalist/supplierCustomer?{}'.format(data.replace('no', 'No').replace('ty', 'Ty').replace('index', 'Index')), headers=self.normal_headers,timeout=5).json()
                    now_res = 3
                except:
                    if self.c == 3:
                        self.c = 0
                    else:
                        self.c +=1
                    self.normal_headers['Cookie'] = self.cookie_pool[self.c]
                    print('重连')
                    now_res+=1

            if res == None:
                time.sleep(3)
                print('没有供应商信息')
                break
            if res['data'] == []:
                print('获取完毕')
                break
            for d in res['data']:
                KeyNo = d['KeyNo']
                CompanyName = d['CompanyName']
                Proportion = d['Proportion']
                Quota = d['Quota']
                DateTime = str(d['ReportYear']) + '-' + str(d['Month'])
                Source = d['Source']
                Relationship = d['Relationship']
                self.save_gys(KeyNo,CompanyName,Proportion,Quota,DateTime,Source,Relationship)
            time.sleep(1)
            page += 1
            self.c += 1
        return True

    def save_zlxx(self,xh,Title,KindCodeDesc,LegalStatusDesc,ApplicationNumber,ApplicationDate,PublicationNumber,PublicationDate,InventorList):
        ApplicationDate = time.strftime("%Y-%m-%d", time.localtime(int(ApplicationDate)))
        PublicationDate = time.strftime("%Y-%m-%d", time.localtime(int(PublicationDate)))
        self.ws.cell(self.row_,10,xh)
        self.ws.cell(self.row_, 11, Title)
        self.ws.cell(self.row_, 12, KindCodeDesc)
        self.ws.cell(self.row_, 13, LegalStatusDesc)
        self.ws.cell(self.row_, 14, ApplicationNumber)
        self.ws.cell(self.row_, 15, ApplicationDate)
        self.ws.cell(self.row_, 16, PublicationNumber)
        self.ws.cell(self.row_, 17, PublicationDate)
        self.ws.cell(self.row_, 18, ','.join(InventorList))
        self.row_+=1
    def save_gys(self,KeyNo,CompanyName,Proportion,Quota,DateTime,Source,Relationship):
        self.ws.cell(self.row,3,KeyNo)
        self.ws.cell(self.row, 4, CompanyName)
        self.ws.cell(self.row, 5, Proportion)
        self.ws.cell(self.row, 6, Quota)
        self.ws.cell(self.row, 7, DateTime)
        self.ws.cell(self.row, 8, Source)
        self.ws.cell(self.row, 9, Relationship)
        self.row+=1

    def run(self):
        no_list = self.load_no()
        for no in no_list:
            no = no.replace('\n', '').strip()
            print(no)
            with open('over.txt','r',encoding='utf-8')as f:
                over_no = f.read()
            if no in over_no:
                continue
            self.ws.cell(self.row,1,no)
            if not self.res_zlxx(no):
                raise Exception('异常')
            if not self.res_gxy(no):
                raise Exception('异常')
            if self.row>self.row_:
                self.row_= self.row
            else:
                self.row = self.row_
            self.row+=1
            self.row_+=1
            time.sleep(2)
            self.wb.save('info.xlsx')
            with open('over.txt','a',encoding='utf-8')as f:
                f.write(no+'\n')

if __name__ == '__main__':
    while True:
        try:
            spider = Spider()
            spider.run()
        except:
            time.sleep(5)
